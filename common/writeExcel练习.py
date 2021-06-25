"""
功能描述：
编写人：SongXuwen
编写日期：
实现逻辑：
    1.
    2.
    3.

"""
import os

import xlrd

from xlutils.copy import copy

# rb = xlrd.open_workbook('data.xls')
# wb = copy(rb)
# ws = wb.get_sheet(0)
# print(ws.name)
# ws.write(1, 1, 'value')
# wb.save('data3.xlsx')


import openpyxl
inwb = openpyxl.load_workbook('data.xls')
sheetnames = inwb.sheetnames # 获取所有的sheet的name
ws = inwb[sheetnames[0]]
rows = ws.max_row
cols = ws.max_column
outwb = openpyxl.Workbook()
outws = outwb.create_sheet(index=2)
for row in range(1, 100):
    for col in range(1,8):
            outws.cell(row,col).value = row*2
    print(row)
saveExcel = 'data2.xlsx'
outwb.save(saveExcel)